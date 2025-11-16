from openpyxl import load_workbook
import csv

wb = load_workbook("doc/imports/base.xlsx", data_only=True)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    with open(f"{sheet}.csv", "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
