"""Import helper that loads data from an Excel workbook."""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Dict, List
import sys

from openpyxl import load_workbook

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.append(str(ROOT_DIR))

from scripts import import_utils


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Lee un Excel y sincroniza los datos con Supabase"
    )
    parser.add_argument(
        "--file",
        default="doc/imports/base.xlsx",
        help="Ruta del archivo .xlsx a procesar",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo muestra cuántos registros se importarían",
    )
    return parser.parse_args()


def normalize(value: Any) -> Any:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8")
    return value


def sheet_to_dicts(sheet) -> List[Dict[str, Any]]:
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx == 0:
            for col_index, header in enumerate(row):
                header_name = str(header).strip() if header else f"column_{col_index}"
                headers.append(header_name)
            continue
        record: Dict[str, Any] = {}
        for header, value in zip(headers, row):
            if value is None or header is None:
                continue
            record[header] = normalize(value)
        if record:
            rows.append(record)
    return rows


def main() -> None:
    args = parse_args()
    excel_path = Path(args.file)
    if not excel_path.exists():
        raise SystemExit(f"No se encontró el archivo {excel_path}")

    workbook = load_workbook(excel_path, data_only=True)
    summary = {}
    for table in import_utils.TABLE_SEQUENCE:
        if table not in workbook.sheetnames:
            summary[table] = {"rows": 0, "imported": 0}
            continue
        sheet = workbook[table]
        rows = sheet_to_dicts(sheet)
        if table == "artifact_links":
            rows = import_utils.deduplicate_links(rows)
        inserted = import_utils.upsert_rows(table, rows, dry_run=args.dry_run)
        summary[table] = {"rows": len(rows), "imported": inserted}

    for table, stats in summary.items():
        status = "SKIPPED" if stats["rows"] == 0 else "OK"
        print(f"[{status}] {table}: {stats['imported']} filas importadas")


if __name__ == "__main__":
    main()
